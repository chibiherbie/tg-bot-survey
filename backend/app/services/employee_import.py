from __future__ import annotations

import io
import json
from collections.abc import Iterable
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from repositories.checklist import EmployeeRepository, PositionRepository
from services.app_settings import AppSettingsService
from services.base import BaseService

CONFIG_PATH = Path(__file__).with_name("employee_import_config.json")
IMPORT_CONFIG_KEY = "employee_import_config"


@dataclass(slots=True)
class ImportConfig:
    sheet_name: str | None
    column_tab_number: str
    column_position: str

    @classmethod
    def from_mapping(cls, data: dict[str, Any]) -> ImportConfig:
        columns = data.get("columns") or {}
        required_fields = {"tab_number", "position"}
        missing = [field for field in required_fields if field not in columns]
        if missing:
            raise ValueError(
                "Config missing columns: " + ", ".join(missing),
            )
        return cls(
            sheet_name=data.get("sheet_name"),
            column_tab_number=str(columns["tab_number"]).strip(),
            column_position=str(columns["position"]).strip(),
        )


@dataclass(slots=True)
class ImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    deactivated: int = 0

    def as_message(self) -> str:
        return (
            f"Создано: {self.created}\n"
            f"Обновлено: {self.updated}\n"
            f"Пропущено (дубликаты): {self.skipped}\n"
            f"Деактивировано: {self.deactivated}"
        )


@dataclass(slots=True)
class EmployeeRow:
    tab_number: str
    position_name: str


class EmployeeImportService(BaseService):
    def __init__(
        self,
        position_repo: PositionRepository,
        employee_repo: EmployeeRepository,
        app_settings_service: AppSettingsService,
    ) -> None:
        self.position_repo = position_repo
        self.employee_repo = employee_repo
        self.app_settings_service = app_settings_service

    async def import_from_bytes(
        self,
        data: bytes,
        *,
        config_path: Path | None = None,
        sheet_name: str | None = None,
    ) -> ImportStats:
        config = await self._load_config(config_path)
        if sheet_name is not None:
            config = replace(config, sheet_name=sheet_name)
        rows = list(self._read_rows(io.BytesIO(data), config))
        return await self._process_rows(rows)

    async def import_from_path(
        self,
        path: Path,
        *,
        config_path: Path | None = None,
        sheet_name: str | None = None,
    ) -> ImportStats:
        with path.open("rb") as fp:
            data = fp.read()
        return await self.import_from_bytes(
            data,
            config_path=config_path,
            sheet_name=sheet_name,
        )

    async def _process_rows(self, rows: Iterable[EmployeeRow]) -> ImportStats:
        stats = ImportStats()
        processed_tab_numbers: set[str] = set()

        for row in rows:
            if row.tab_number in processed_tab_numbers:
                stats.skipped += 1
                continue
            processed_tab_numbers.add(row.tab_number)

            position = await self._ensure_position(row.position_name)
            existing = await self.employee_repo.get_by_tab_number(
                row.tab_number,
            )
            payload = {
                "position_id": position.id,
                "is_active": True,
            }
            if existing:
                await self.employee_repo.update(existing, payload)
                stats.updated += 1
            else:
                payload["tab_number"] = row.tab_number
                await self.employee_repo.create(payload)
                stats.created += 1
        stats.deactivated += await self._deactivate_missing(
            processed_tab_numbers,
        )
        return stats

    async def _ensure_position(self, name: str):
        if existing := await self.position_repo.get_by_name(name):
            return existing
        return await self.position_repo.create({"name": name})

    async def _deactivate_missing(
        self,
        processed_tab_numbers: set[str],
    ) -> int:
        deactivated = 0
        employees = await self.employee_repo.list_all()
        for employee in employees:
            if employee.tab_number in processed_tab_numbers:
                continue
            if employee.is_active:
                await self.employee_repo.update(employee, {"is_active": False})
                deactivated += 1
        return deactivated

    def _read_rows(
        self,
        buffer: io.BytesIO,
        config: ImportConfig,
    ) -> Iterable[EmployeeRow]:
        workbook = load_workbook(
            filename=buffer,
            data_only=True,
            read_only=True,
        )
        try:
            sheet = self._select_sheet(workbook, config.sheet_name)
            rows = sheet.iter_rows(values_only=True)
            headers = self._extract_headers(rows)

            tab_idx = self._resolve_column(headers, config.column_tab_number)
            position_idx = self._resolve_column(
                headers,
                config.column_position,
            )

            for row in rows:
                tab_value = self._normalize_tab_number(
                    row[tab_idx] if tab_idx < len(row) else None,
                )
                if not tab_value:
                    continue
                position_value = self._normalize_string(
                    row[position_idx] if position_idx < len(row) else None,
                )
                if not position_value:
                    continue
                yield EmployeeRow(
                    tab_number=tab_value,
                    position_name=position_value,
                )
        finally:
            workbook.close()

    @staticmethod
    def _select_sheet(workbook, sheet_name: str | None):
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Лист '{sheet_name}' не найден в файле")
            return workbook[sheet_name]
        return workbook.active

    @staticmethod
    def _extract_headers(rows) -> list[str]:
        try:
            headers = next(rows)
        except StopIteration as exc:
            raise ValueError("Файл пустой") from exc
        return [
            str(value).strip() if value is not None else ""
            for value in headers
        ]

    @staticmethod
    def _resolve_column(headers: list[str], column_name: str) -> int:
        try:
            return headers.index(column_name)
        except ValueError as exc:
            raise ValueError(
                f"Колонка '{column_name}' не найдена в файле",
            ) from exc

    @staticmethod
    def _normalize_tab_number(value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned or None
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            return str(int(value))
        return str(value).strip() or None

    @staticmethod
    def _normalize_string(value: Any) -> str | None:
        if value is None:
            return None
        cleaned = str(value).strip()
        return cleaned or None

    async def _load_config(self, config_path: Path | None) -> ImportConfig:
        if config_path is not None:
            with config_path.open("r", encoding="utf-8") as fp:
                raw = json.load(fp)
            return ImportConfig.from_mapping(raw)

        settings_value = await self.app_settings_service.get_json(
            IMPORT_CONFIG_KEY,
        )
        if settings_value:
            return ImportConfig.from_mapping(settings_value)

        if CONFIG_PATH.exists():
            with CONFIG_PATH.open("r", encoding="utf-8") as fp:
                raw = json.load(fp)
            return ImportConfig.from_mapping(raw)

        return ImportConfig(
            sheet_name=None,
            column_tab_number="Табельный номер",
            column_position="Должность",
        )


__all__ = [
    "EmployeeImportService",
    "ImportStats",
]
